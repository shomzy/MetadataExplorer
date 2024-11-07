import glob
import json
import os
import re
import pandas as pd
import win32com.client as win32
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font


def merge_objects(obj1, obj2):
    if isinstance(obj1, tuple):
        o1 = obj1
    else:
        o1 = (obj1,)
    if isinstance(obj2, tuple):
        o2 = obj2
    else:
        o2 = (obj2,)

    return o1 + o2


def write_repository_paths_data_to_excel(data, filename, delimiter):
    filename = os.path.abspath(filename)
    # Create a DataFrame from the data and add the new column
    for entry in data:
        entry['Used by Lunaris'] = f"False"

        if entry['values_file_path'] is not None:
            last_dir = os.path.basename(os.path.dirname(entry['values_file_path']))  # 'leaves'
            file_name = os.path.basename(entry['values_file_path'])

            # Construct the new relative path
            entry['values_file_path'] = f'{last_dir}\\{file_name}'

    df = pd.DataFrame(data)

    # Write the DataFrame to an Excel file
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write the general tab
        df.to_excel(writer, sheet_name='General', index=False)

    # Load the workbook to add hyperlinks
    workbook = load_workbook(filename)
    general_sheet = workbook['General']

    # Add hyperlinks to the 'values' column
    for row in general_sheet.iter_rows(min_row=2, min_col=1, max_col=len(df.columns)):
        for cell in row:
            if cell.column == 3:  # Assuming 'values' is in the third column
                if cell.value is not None:
                    cell.hyperlink = None
                    cell.font = Font(color="0000FF", underline="single")  # Blue color and underline for hyperlink

    workbook.save(filename)

    source_filename = filename
    target_filename = change_file_extension(filename, '.xlsm')
    convert_to_xlsm(source_filename, target_filename, column_pos=3, delimiter=delimiter)


def convert_to_xlsm(source_filename, target_filename, column_pos, delimiter):
    """Converts a .xlsx file to .xlsm format using win32com.client."""
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    workbook = excel.Workbooks.Open(source_filename)
    excel.Visible = False

    # Add the VBA event handler to the workbook and worksheet
    excel_module = workbook.VBProject.VBComponents(workbook.CodeName)  # For ThisWorkbook module
    sheet_module = workbook.VBProject.VBComponents(workbook.Worksheets(1).CodeName)  # For specific sheet module

    # VBA code to set drive and directory when the workbook is opened
    workbook_vba_code = '''
    Private Sub Workbook_Open()
        ' Set the drive and working directory once when the workbook is opened
        Call SetDriveAndDirectory
    End Sub

    Private Sub SetDriveAndDirectory()
        ' Change the current drive to the workbook's drive
        On Error Resume Next
        ChDrive ThisWorkbook.Path
        On Error GoTo 0

        ' Change the current directory to the workbook's directory
        ChDir ThisWorkbook.Path
    End Sub
    '''

    # VBA code to handle selection change and open CSV files directly
    sheet_vba_code = f'''
    Private Sub Worksheet_SelectionChange(ByVal Target As Range)
        ' Check if the selected cell is in the specified column (e.g., column 3)
        If Target.Column = {column_pos} Then
            Dim relativeFilePath As String
            relativeFilePath = Target.Value ' Assuming the cell value is the relative file path

            ' Ensure the cell contains a value
            If Len(relativeFilePath) > 0 Then
                ' Combine the workbook path with the relative file path
                Dim fullPath As String
                fullPath = ThisWorkbook.Path & Application.PathSeparator & relativeFilePath

                ' Check if the file exists
                If Dir(fullPath) <> "" Then
                    ' Disable events to prevent recursive calls
                    ' Application.EnableEvents = False

                    ' Open the CSV file with the specified delimiter in Excel
                    Application.Workbooks.OpenText _
                        Filename:=fullPath, _
                        Origin:=xlMSDOS, _
                        StartRow:=1, _
                        DataType:=xlDelimited, _
                        TextQualifier:=xlDoubleQuote, _
                        ConsecutiveDelimiter:=False, _
                        Tab:=False, _
                        Semicolon:=False, _
                        Comma:=False, _
                        Space:=False, _
                        Other:=True, _
                        OtherChar:="{delimiter}", _
                        FieldInfo:=Array(1, 1), _
                        TrailingMinusNumbers:=True

                    ' Re-enable events after the file is opened
                    ' Application.EnableEvents = True
                Else
                    MsgBox "File not found: " & fullPath
                End If
            End If
        End If
    End Sub
    '''

    # Add VBA code to the ThisWorkbook module
    excel_module.CodeModule.AddFromString(workbook_vba_code)

    # Add VBA code to the worksheet module
    sheet_module.CodeModule.AddFromString(sheet_vba_code)

    # Save as .xlsm
    workbook.SaveAs(Filename=target_filename, FileFormat=52)  # 52 is the code for .xlsm format
    workbook.Close(False)
    excel.Quit()

    # Remove the temporary .xlsx file
    os.remove(source_filename)


def change_file_extension(file_path, new_extension):
    # Split the file path into the root and extension
    root, _ = os.path.splitext(file_path)

    # Create the new file path with the desired extension
    new_file_path = root + new_extension

    return new_file_path


def delete_all_files(dir_path):
    files = glob.glob(os.path.join(dir_path, '*'))

    for file_path in files:
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
        except Exception as e:
            print(f"Error deleting file {file_path}: {e}")


def copy_to_new_stats_instance(source_stats):
    stats = {}
    for k in source_stats:
        stats[k] = {}

    accumulate_stats(stats, source_stats)
    return stats


def accumulate_stats(target_stats, source_stats):
    for k in source_stats:
        for rk in source_stats[k]:
            if k not in target_stats:
                target_stats[k] = {}
            if rk not in target_stats[k]:
                target_stats[k][rk] = source_stats[k][rk]
            else:
                target_stats[k][rk] = target_stats[k][rk] + source_stats[k][rk]


def extract_repo_name_parts(text):
    # Use regular expression to find the text within parentheses and the remaining text
    match = re.match(r'\((.*?)\)\s*(.*)', text)
    if match:
        part1 = match.group(1)  # Text within parentheses
        part2 = match.group(2)  # Remaining text
        return part1, part2
    else:
        return None, None


def is_value_missing(value):
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True

    return False


def flush_value_to_file(agg_leaf, delimiter):
    if agg_leaf.get_leaves_len == 0:
        return

    custom_string='<%%C%%>'

    lines = []
    for lf in agg_leaf.get_leaves():
        repo_type = replace_delimiter_with_custom_string(lf.repo_type, delimiter=delimiter, custom_string=custom_string)
        repo_type = sanitize_for_excel(repo_type)
        repo_name = replace_delimiter_with_custom_string(lf.repo_name, delimiter=delimiter, custom_string=custom_string)
        repo_name = sanitize_for_excel(repo_name)
        repo_item_name = replace_delimiter_with_custom_string(lf.repo_item_name, delimiter=delimiter, custom_string=custom_string)
        repo_item_name = sanitize_for_excel(repo_item_name)
        value = replace_delimiter_with_custom_string(lf.value, delimiter=delimiter, custom_string=custom_string)
        value = sanitize_for_excel(value)

        last_dir = os.path.basename(os.path.dirname(lf.file_path))
        file_name = os.path.basename(lf.file_path)

        # Construct the new relative path
        # leaf_file_path = f'.\\{last_dir}\\{file_name}'
        leaf_file_path = f'=HYPERLINK("..\\{last_dir}\\{file_name}", "Open")'

        lines.append(delimiter.join([repo_type, repo_name, repo_item_name, leaf_file_path, value]))

    with open(agg_leaf.leaf_file_path, 'a', encoding='utf-8') as file:
        file.write("\n".join(lines) + "\n")


def sanitize_for_excel(input_string):
    # Remove control characters and other non-printable characters
    sanitized_string = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', input_string)
    return sanitized_string


def replace_delimiter_with_custom_string(input_string, delimiter, custom_string):
    return input_string.replace(delimiter, custom_string)


def write_json(file_path, data):
    with open(file_path, "w") as file:
        json.dump(data, file)
